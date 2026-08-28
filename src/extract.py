"""업로드된 메일·첨부에서 텍스트를 뽑아 Gemini에 넘긴다."""

from __future__ import annotations

import io
import zipfile
from xml.etree import ElementTree


def extract_upload_text(filename: str, data: bytes) -> str:
    name = str(filename or "").lower()
    if not data:
        return ""
    if name.endswith((".txt", ".csv", ".md", ".eml", ".log", ".json", ".xml")):
        return _decode(data)
    if name.endswith((".xlsx", ".xlsm")):
        return _xlsx_text(data)
    if name.endswith(".docx"):
        return _docx_text(data)
    if name.endswith(".pdf"):
        return _pdf_text(data)
    return _decode(data)


def _decode(data: bytes) -> str:
    for encoding in ("utf-8", "cp949", "euc-kr", "utf-16"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="ignore")


def _xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet in workbook.worksheets:
        parts.append(f"[시트 {sheet.title}]")
        for row in sheet.iter_rows(max_row=80, max_col=20, values_only=True):
            cells = [str(value).strip() for value in row if value is not None and str(value).strip()]
            if cells:
                parts.append("\t".join(cells))
    workbook.close()
    return "\n".join(parts)


def _docx_text(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        xml = archive.read("word/document.xml")
    tree = ElementTree.fromstring(xml)
    texts = [
        node.text
        for node in tree.iter()
        if node.text and node.text.strip()
    ]
    return "\n".join(texts)


def _pdf_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "(PDF 텍스트를 읽지 못했습니다. 본문에 붙여 넣거나 TXT로 저장해 주세요.)"
    reader = PdfReader(io.BytesIO(data))
    pages = []
    for page in reader.pages[:15]:
        pages.append(page.extract_text() or "")
    return "\n".join(pages)
