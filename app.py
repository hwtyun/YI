"""용인공장 특근 인원 조사 취합 — 진입점.

실행: 프로젝트 폴더에서
    streamlit run app.py

로그인 화면 함수는 이 파일에 둔다. Cloud가 src.theme 옛 모듈을 남기면
새 이름을 import할 때 ImportError가 나므로 src.theme에서 새 함수를 가져오지 않는다.
"""

from __future__ import annotations

import hashlib
from pathlib import Path

try:
    from boot import bind_project_package
except ImportError:
    import sys

    def bind_project_package() -> Path:
        root = str(Path(__file__).resolve().parent)
        if root not in sys.path:
            sys.path.insert(0, root)
        return Path(root)

bind_project_package()

try:
    from branding import SITE_TITLE
except ImportError:
    SITE_TITLE = "용인공장 특근·본사요청 취합"

import streamlit as st

from src.auth import get_authenticator, sign_in
from src.config import USERS
from src.db import init_db
from src.secrets_util import missing_secret_names, render_secrets_help

st.set_page_config(
    page_title=SITE_TITLE,
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="collapsed",
)
init_db()

_ROOT = Path(__file__).resolve().parent
_LOGO_CANDIDATES = (
    _ROOT / "static" / "atec_ci.png",
    _ROOT / "ATEC 영문_기본형.png",
)
_PAGE_TEXT_CSS = """
<style>
[data-testid="stHeading"],
[data-testid="stHeading"] *,
[data-testid="stCaption"],
[data-testid="stCaption"] *,
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] *,
[data-testid="stWidgetLabel"],
[data-testid="stWidgetLabel"] *,
.stHeading, .stHeading *,
.stCaption, .stCaption *,
h1, h2, h3, h4, h5, h6,
[data-testid="stMarkdown"] p,
[data-testid="stMarkdown"] span,
[data-testid="stMarkdown"] strong,
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] span,
[data-testid="stMarkdownContainer"] strong {
    color: #000000 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #000000 !important;
}
div[class*="st-key-cal_today"] {
    display: none !important;
}
[data-testid="stDownloadButton"] button {
    background-color: #1f4e79 !important;
    border-color: #1f4e79 !important;
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
    opacity: 1 !important;
}
[data-testid="stDownloadButton"] button p,
[data-testid="stDownloadButton"] button span,
[data-testid="stDownloadButton"] button div {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}
.yi-cal-month,
.yi-cal-month * {
    text-align: center !important;
    font-size: 1.7rem !important;
    font-weight: 700 !important;
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    background: transparent !important;
    background-color: transparent !important;
    line-height: 2.4rem !important;
    margin: 0 !important;
}
.yi-tag-meal {
    background: #fff4e5 !important;
    color: #b35c00 !important;
}
.yi-tag-off {
    background: #f4f6f8 !important;
    color: #7a8794 !important;
}
.yi-cell {
    height: 7.6rem !important;
    min-height: 7.6rem !important;
    max-height: 7.6rem !important;
    padding: 4px 2px 2px 2px !important;
    display: flex !important;
    flex-direction: column !important;
    box-sizing: border-box !important;
    overflow: hidden !important;
}
.yi-cell-num { text-align: center; font-size: 0.95rem; margin-bottom: 2px !important; }
.yi-cell-meta {
    flex: 1;
    min-height: 4.4rem;
    display: flex;
    flex-direction: column;
    justify-content: flex-start;
    gap: 2px;
}
.yi-tag {
    display: block !important;
    border-radius: 8px !important;
    padding: 2px 4px !important;
    margin: 0 2px !important;
    font-size: 0.64rem !important;
    font-weight: 600 !important;
    text-align: center !important;
    line-height: 1.3 !important;
    overflow: hidden !important;
}
.yi-tag-spacer { visibility: hidden !important; background: transparent !important; }
.yi-sun, .yi-sun * {
    color: #d94848 !important;
    -webkit-text-fill-color: #d94848 !important;
}
.yi-sat, .yi-sat * {
    color: #1a5fb4 !important;
    -webkit-text-fill-color: #1a5fb4 !important;
}
[data-testid="stMarkdownContainer"]:has(.yi-cell) {
    min-height: 7.6rem !important;
}
[data-testid="stMarkdownContainer"]:has(.yi-cell) p {
    margin: 0 !important;
}
</style>
"""
_LOGIN_CSS = """
<style>
html, body, .stApp, [data-testid="stAppViewContainer"],
[data-testid="stAppViewContainer"] > section,
[data-testid="stMain"], [data-testid="stMainBlockContainer"],
.stMain, section.main, .main, .block-container,
[data-testid="stHeader"], header,
div[class*="st-key-yi_login"] {
    background: #ffffff !important;
    background-color: #ffffff !important;
    color-scheme: light !important;
}
.stApp, [data-testid="stAppViewContainer"] {
    --background-color: #ffffff !important;
    --secondary-background-color: #87CEEB !important;
    --st-background-color: #ffffff !important;
}
div[class*="st-key-yi_login"] h1,
div[class*="st-key-yi_login"] h2,
div[class*="st-key-yi_login"] p,
div[class*="st-key-yi_login"] label,
div[class*="st-key-yi_login"] span,
div[class*="st-key-yi_login"] [data-testid="stWidgetLabel"],
div[class*="st-key-yi_login"] [data-testid="stWidgetLabel"] *,
div[class*="st-key-yi_login"] [data-testid="stCheckbox"] p,
div[class*="st-key-yi_login"] [data-testid="stCheckbox"] span,
div[class*="st-key-yi_login"] [data-testid="stCheckbox"] label,
div[class*="st-key-yi_login"] [data-testid="stMarkdown"] p,
.yi-site-title, .yi-login-caption {
    color: #000000 !important;
    opacity: 1 !important;
    -webkit-text-fill-color: #000000 !important;
}
div[class*="st-key-yi_login"] [data-testid="stFormSubmitButton"] button,
div[class*="st-key-yi_login"] [data-testid="stFormSubmitButton"] p,
div[class*="st-key-yi_login"] [data-testid="stFormSubmitButton"] span {
    color: #ffffff !important;
    -webkit-text-fill-color: #ffffff !important;
}
div[class*="st-key-yi_login"] [data-testid="stTextInput"] [data-baseweb="input"],
div[class*="st-key-yi_login"] [data-testid="stTextInput"] [data-baseweb="base-input"],
div[class*="st-key-yi_login"] [data-testid="stTextInput"] input,
div[class*="st-key-yi_login"] [data-testid="stTextInput"] > div > div {
    background: #87CEEB !important;
    background-color: #87CEEB !important;
    color: #000000 !important;
    -webkit-text-fill-color: #000000 !important;
    caret-color: #000000 !important;
}
div[class*="st-key-yi_login"] [data-testid="stTextInput"] button,
div[class*="st-key-yi_login"] [data-testid="stTextInput"] svg {
    color: #000000 !important;
    fill: #000000 !important;
}
</style>
"""


def _inject_css(css: str) -> None:
    try:
        st.html(css)
    except Exception:
        st.markdown(css, unsafe_allow_html=True)


def _inject_login_styles() -> None:
    _inject_css(_PAGE_TEXT_CSS)
    _inject_css(_LOGIN_CSS)


def _render_site_title() -> None:
    st.markdown(
        f'<h1 class="yi-site-title" style="color:#000000 !important;'
        f'-webkit-text-fill-color:#000000 !important;opacity:1 !important;'
        f'font-weight:700;font-size:2rem;margin:0.15rem 0 0.4rem 0;">'
        f"{SITE_TITLE}</h1>",
        unsafe_allow_html=True,
    )


def _render_login_caption() -> None:
    st.markdown(
        '<p class="yi-login-caption" style="color:#000000 !important;'
        '-webkit-text-fill-color:#000000 !important;opacity:1 !important;'
        'font-size:0.95rem;margin:0 0 0.7rem 0;">'
        "아이디로 로그인한 뒤 본인 화면만 사용합니다. 휴대폰에서도 입력할 수 있습니다.</p>",
        unsafe_allow_html=True,
    )


def _render_login_logo(width: int = 160) -> None:
    for path in _LOGO_CANDIDATES:
        if path.exists():
            st.image(str(path), width=width)
            return
    st.markdown('<div style="font-weight:700;color:#9b1c2e">ATEC</div>', unsafe_allow_html=True)


def _secrets_ready() -> bool:
    missing = missing_secret_names()
    if not missing:
        return True
    _inject_login_styles()
    _render_login_logo(160)
    _render_site_title()
    render_secrets_help(missing)
    return False


def _authenticator():
    expiry = 30.0 if st.session_state.get("remember_login", True) else 0.0
    authenticator = get_authenticator(expiry)
    st.session_state["authenticator"] = authenticator
    return authenticator


def _restore_cookie(authenticator) -> None:
    if st.session_state.get("authentication_status"):
        return
    try:
        token = authenticator.cookie_controller.get_cookie()
    except Exception:
        return
    if token:
        authenticator.authentication_controller.login(token=token)


def _render_login(authenticator) -> None:
    _inject_login_styles()
    with st.container(key="yi_login"):
        _render_login_logo(160)
        _render_site_title()
        _render_login_caption()
        st.checkbox("로그인 상태 유지", value=True, key="remember_login")

        with st.form("yi_factory_login"):
            username = st.text_input("아이디", key="login_username")
            password = st.text_input("비밀번호", type="password", key="login_password")
            submitted = st.form_submit_button("로그인", type="primary")

    if submitted:
        if sign_in(authenticator, username, password):
            try:
                authenticator.cookie_controller.set_cookie()
            except Exception:
                pass
            st.rerun()
            return
        st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
        return
    if st.session_state.get("authentication_status") is False:
        st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
        return
    st.info("아이디와 비밀번호를 입력한 뒤 로그인하세요.")


def _parse_roster_bytes(data: bytes) -> dict:
    """성명·회사·팀·고용형태 양식. Cloud의 옛 src.config와 무관하게 동작한다."""
    from io import BytesIO

    from openpyxl import load_workbook

    def fold(value: str) -> str:
        return str(value or "").strip().replace(" ", "").replace("-", "").lower()

    company_map = {
        "에이텍모빌리티": "에이텍모빌리티",
        "모빌리티": "에이텍모빌리티",
        "aitechmobility": "에이텍모빌리티",
        "atecmobility": "에이텍모빌리티",
        "에이텍컴퓨터": "에이텍컴퓨터",
        "컴퓨터": "에이텍컴퓨터",
        "aitechcomputer": "에이텍컴퓨터",
        "ateccomputer": "에이텍컴퓨터",
    }
    team_map = {
        "구매팀": "구매팀",
        "구매": "구매팀",
        "제조팀": "제조팀",
        "제조": "제조팀",
        "생산관리팀": "생산관리팀",
        "생산관리": "생산관리팀",
        "용인공장": "생산관리팀",
        "공장": "생산관리팀",
        "품질보증팀": "품질보증팀",
        "품질보증": "품질보증팀",
        "품질팀": "품질보증팀",
        "생산기술파트": "생산기술파트",
        "생산기술": "생산기술파트",
        "생산기술팀": "생산기술파트",
        "자재파트": "자재파트",
        "자재": "자재파트",
        "자재팀": "자재파트",
    }
    emp_map = {
        "정규직": "정규직",
        "정규": "정규직",
        "계약직": "계약직",
        "계약": "계약직",
        "일용직": "일용직",
        "일용": "일용직",
    }

    def as_company(value: str) -> str | None:
        text = str(value or "").strip()
        if text in ("에이텍모빌리티", "에이텍컴퓨터"):
            return text
        return company_map.get(fold(text))

    def as_team(value: str) -> str | None:
        text = str(value or "").strip()
        if text in team_map.values():
            return text
        return team_map.get(fold(text)) or team_map.get(text)

    def as_employment(value: str) -> str:
        text = str(value or "").strip()
        if text in ("정규직", "계약직", "일용직"):
            return text
        mapped = emp_map.get(fold(text))
        return mapped or "정규직"

    workbook = load_workbook(BytesIO(data), data_only=True)
    sheet = workbook.active
    rows: list[dict] = []
    errors: list[str] = []
    header = None
    start = 2
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, max_col=8, values_only=True), start=1):
        texts = [str(item or "").strip() for item in row]
        if "성명" in texts and "회사" in texts and "팀" in texts:
            header = texts
            start = index + 1
            break
    if header is None:
        return {"rows": [], "errors": ["성명·회사·팀 헤더를 찾지 못했습니다."]}
    name_col = header.index("성명")
    company_col = header.index("회사") if "회사" in header else 1
    team_col = header.index("팀") if "팀" in header else 2
    type_col = header.index("고용형태") if "고용형태" in header else 3
    seen: set[tuple[str, str, str]] = set()
    for excel_row, row in enumerate(sheet.iter_rows(min_row=start, max_col=8, values_only=True), start=start):
        values = list(row)

        def cell(idx: int) -> str:
            if idx >= len(values) or values[idx] is None:
                return ""
            return str(values[idx]).strip()

        name = cell(name_col)
        if not name or name == "안내" or name.startswith(("회사는", "팀은", "고용형태는", "자주", "고정")):
            continue
        company = as_company(cell(company_col))
        team_raw = cell(team_col)
        team = as_team(team_raw)
        employment = as_employment(cell(type_col))
        if company is None:
            errors.append(f"{excel_row}행 '{name}': 회사는 에이텍모빌리티 또는 에이텍컴퓨터여야 합니다.")
            continue
        if team is None:
            errors.append(f"{excel_row}행 '{name}': 팀을 확인할 수 없습니다. ({team_raw})")
            continue
        key = (name, company, team)
        if key in seen:
            errors.append(f"{excel_row}행 '{name}': 같은 회사·팀에 이름이 중복됩니다.")
            continue
        seen.add(key)
        rows.append(
            {"name": name, "company": company, "team": team, "employment_type": employment}
        )
    if not rows and not errors:
        errors.append("유효한 임직원 행이 없습니다.")
    return {"rows": rows, "errors": errors}


def _render_prodadmin_excel_roster(username: str) -> None:
    """Cloud가 옛 roster 화면만 읽어도, app.py만 올리면 엑셀 받기·올리기가 보이게 한다."""
    from src.config import COMPANIES, EMPLOYMENT_TYPES, ROLE_ADMIN, SUBMITTING_TEAMS, primary_role
    from src.excel_io import build_employee_template
    from src.store import AccessDenied, list_employees, replace_employee_roster

    if primary_role(username) != ROLE_ADMIN:
        return

    st.subheader("재직인원 명부")
    st.caption(
        "양식에 성명·회사·팀·고용형태를 적은 뒤 아래에서 엑셀만 올리면 바로 반영됩니다. "
        "고용형태에 직급(이사·책임·선임·사원 등)을 적어도 정규직으로 저장합니다."
    )
    current = list_employees(username)
    left, right = st.columns(2)
    with left:
        st.download_button(
            "재직인원 양식 다운받기",
            data=build_employee_template([]),
            file_name="재직인원_양식.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="app_dl_roster_template",
        )
    with right:
        st.download_button(
            "현재 재직인원 엑셀 받기",
            data=build_employee_template(current),
            file_name="재직인원_전체.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="app_dl_roster_current",
            disabled=not current,
        )

    st.session_state["_yi_excel_roster_shown"] = True

    uploaded = st.file_uploader("재직인원 엑셀 업로드", type=["xlsx"], key="app_up_roster")
    if uploaded is not None:
        raw = uploaded.getvalue()
        parsed = _parse_roster_bytes(raw)
        if parsed["errors"]:
            st.error("일부 행은 건너뛰었습니다.\n" + "\n".join(f"- {item}" for item in parsed["errors"]))
        rows = parsed["rows"]
        if rows:
            sig = hashlib.sha256(raw).hexdigest()
            if st.session_state.get("app_roster_sig") != sig:
                try:
                    count = replace_employee_roster(username, rows)
                    st.session_state["app_roster_sig"] = sig
                    st.success(f"재직인원 {count}명을 반영했습니다.")
                    st.rerun()
                except AccessDenied as exc:
                    st.error(str(exc))
            else:
                st.success(f"반영된 인원 {len(rows)}명")
                st.dataframe(
                    [
                        {
                            "성명": item["name"],
                            "회사": item["company"],
                            "팀": item["team"],
                            "고용형태": item["employment_type"],
                        }
                        for item in rows
                    ],
                    hide_index=True,
                    width="stretch",
                )
        elif not parsed["errors"]:
            st.warning("파일에서 인원을 읽지 못했습니다.")

    st.markdown("**현재 전체 재직인원**")
    if not current:
        st.info(
            "아직 명부가 없습니다. 「재직인원 양식 다운받기」로 양식을 받아 올린 뒤 확인하세요. "
            f"회사: {' · '.join(COMPANIES)} / 팀: {' · '.join(SUBMITTING_TEAMS)} / 고용형태: {' · '.join(EMPLOYMENT_TYPES)}"
        )
    else:
        st.caption(f"전체 {len(current)}명")
        st.dataframe(
            [
                {
                    "성명": item["name"],
                    "회사": item["company"],
                    "팀": item["team"],
                    "고용형태": item["employment_type"],
                }
                for item in current
            ],
            hide_index=True,
            width="stretch",
        )


def _shift_month_app(value, delta: int):
    from datetime import date

    month = value.month - 1 + delta
    year = value.year + month // 12
    month = month % 12 + 1
    return date(year, month, 1)


def _nav_month_button(label: str, key: str) -> bool:
    try:
        return st.button(label, key=key, width="stretch")
    except TypeError:
        return st.button(label, key=key, use_container_width=True)


def _render_calendar_month_nav(month_cursor) -> None:
    """이전달 · 2026년 8월 · 다음달. app.py에 두어 Cloud가 src 옛 파일을 써도 반영된다."""
    try:
        prev_col, month_col, next_col = st.columns([1, 1.7, 1], vertical_alignment="center")
    except TypeError:
        prev_col, month_col, next_col = st.columns([1, 1.7, 1])
    with prev_col:
        if _nav_month_button("이전달", "cal_prev"):
            st.session_state["cal_month"] = _shift_month_app(month_cursor, -1)
            st.rerun()
    with month_col:
        label = f"{month_cursor.year}년 {month_cursor.month}월"
        html = (
            f"<div style='text-align:center;font-size:28px;font-weight:800;"
            f"color:#111111;-webkit-text-fill-color:#111111;opacity:1;"
            f"line-height:40px;margin:0;background:transparent;'>{label}</div>"
        )
        try:
            st.html(html)
        except Exception:
            st.markdown(
                f"<h2 style='text-align:center;margin:0.15rem 0;color:#000000;"
                f"-webkit-text-fill-color:#000000;font-size:28px;font-weight:800;'>"
                f"{label}</h2>",
                unsafe_allow_html=True,
            )
    with next_col:
        if _nav_month_button("다음달", "cal_next"):
            st.session_state["cal_month"] = _shift_month_app(month_cursor, 1)
            st.rerun()


def _render_team_day_excel_io(username: str, team: str, survey: dict, work_date: str) -> None:
    """해당 팀 명단 엑셀 받기·올리기. Cloud가 옛 survey_editor를 써도 app.py만 올리면 동작한다."""
    from io import BytesIO

    from openpyxl import Workbook, load_workbook

    from src.store import (
        AccessDenied,
        enrich_entry_from_roster,
        list_employees,
        list_entries,
        replace_team_entries,
        survey_edit_status,
    )

    try:
        from src.excel_io import build_team_input_template, parse_team_input_workbook
    except ImportError:
        build_team_input_template = None
        parse_team_input_workbook = None

    survey_id = int(survey["id"])
    can_edit, reason = survey_edit_status(username, survey_id)
    roster = list_employees(username, team)
    existing = [
        item
        for item in list_entries(username, survey_id)
        if str(item.get("team") or "") == team
    ]
    saved = [item for item in existing if str(item.get("work_date") or "") == work_date]

    def build_local() -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "특근입력"
        sheet.append(["팀", team, "특근일자", work_date])
        sheet.append([])
        sheet.append(["회사", "성명", "고용형태", "특근", "근무시간", "식수인원", "비고"])
        saved_by_name = {
            (str(item.get("name") or "").strip(), str(item.get("company") or "").strip()): item
            for item in saved
            if not item.get("is_manual")
        }
        for employee in roster:
            key = (str(employee.get("name") or "").strip(), str(employee.get("company") or "").strip())
            found = saved_by_name.get(key)
            sheet.append(
                [
                    employee.get("company") or "",
                    employee.get("name") or "",
                    employee.get("employment_type") or "",
                    "예" if found else "",
                    found.get("work_hours") if found and found.get("work_hours") is not None else 8,
                    found.get("meal_count") if found and found.get("meal_count") is not None else 1,
                    (found.get("note") or "") if found else "",
                ]
            )
        for item in saved:
            if not item.get("is_manual"):
                continue
            sheet.append(
                [
                    item.get("company") or "",
                    item.get("name") or "",
                    "일용직",
                    "예",
                    item.get("work_hours") if item.get("work_hours") is not None else 8,
                    item.get("meal_count") if item.get("meal_count") is not None else 1,
                    item.get("note") or "",
                ]
            )
        for _ in range(8):
            sheet.append(["", "", "일용직", "", "", "", ""])
        sheet.append([])
        sheet.append(["안내"])
        sheet.append(["특근하는 사람만 특근 칸에 예 를 적으세요. 비우면 특근이 아닙니다."])
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def parse_local(raw: bytes) -> dict:
        yes = {"예", "y", "yes", "1", "true", "o", "ㅇ", "체크", "특근", "v"}
        workbook = load_workbook(BytesIO(raw), data_only=True)
        sheet = workbook.active
        header = None
        start = 2
        for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, max_col=10, values_only=True), start=1):
            texts = [str(item or "").strip() for item in row]
            if "성명" in texts and "특근" in texts:
                header = texts
                start = index + 1
                break
        if header is None:
            return {"entries": [], "errors": ["성명·특근 헤더를 찾지 못했습니다."]}
        name_col = header.index("성명")
        company_col = header.index("회사") if "회사" in header else 0
        type_col = header.index("고용형태") if "고용형태" in header else 2
        mark_col = header.index("특근") if "특근" in header else 3
        hours_col = header.index("근무시간") if "근무시간" in header else 4
        meal_col = header.index("식수인원") if "식수인원" in header else 5
        note_col = header.index("비고") if "비고" in header else 6
        entries = []
        errors = []
        for excel_row, row in enumerate(sheet.iter_rows(min_row=start, max_col=10, values_only=True), start=start):
            values = list(row)

            def cell(idx: int):
                if idx >= len(values):
                    return ""
                return values[idx]

            name = str(cell(name_col) or "").strip()
            if not name or name == "안내" or name.startswith("특근하는"):
                continue
            mark = cell(mark_col)
            mark_text = str(mark or "").strip().replace(" ", "").lower()
            if mark is not True and mark_text not in yes and not mark_text.startswith("예"):
                continue
            hours_raw = str(cell(hours_col) or "").replace("H", "").replace("h", "").strip()
            try:
                hours = float(hours_raw) if hours_raw else 8.0
            except ValueError:
                hours = 8.0
            if hours <= 0:
                hours = 8.0
            meals_raw = str(cell(meal_col) or "").strip()
            try:
                meals = float(meals_raw) if meals_raw else 1.0
            except ValueError:
                meals = 1.0
            entries.append(
                {
                    "seq_no": len(entries) + 1,
                    "name": name,
                    "company": str(cell(company_col) or "").strip() or None,
                    "employment_type": str(cell(type_col) or "").strip() or None,
                    "work_date": work_date,
                    "work_hours": hours,
                    "meal_count": meals,
                    "note": str(cell(note_col) or "").strip() or None,
                    "is_manual": 0,
                    "team": team,
                }
            )
        if not entries:
            errors.append("특근 칸에 예를 적은 인원이 없습니다.")
        return {"entries": entries, "errors": errors}

    data = (
        build_team_input_template(team, work_date, roster, saved)
        if build_team_input_template is not None
        else build_local()
    )
    st.caption(
        "양식에 지금 팀 재직인원이 들어 있습니다. "
        "특근하는 사람만 특근 칸에 예를 적고 시간·식수를 채운 뒤 올리세요."
    )
    down_col, up_col = st.columns(2)
    with down_col:
        st.download_button(
            "우리 팀 명단 엑셀 받기",
            data=data,
            file_name=f"{team}_{work_date}_특근입력.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"app_dl_team_{survey_id}_{team}_{work_date}",
        )
    with up_col:
        uploaded = st.file_uploader(
            "작성한 엑셀 올리기",
            type=["xlsx"],
            key=f"app_up_team_{survey_id}_{team}_{work_date}",
        )
    if uploaded is None:
        return
    raw = uploaded.getvalue()
    if parse_team_input_workbook is not None:
        parsed_obj = parse_team_input_workbook(raw, default_team=team, default_date=work_date)
        parsed = {"entries": parsed_obj.entries, "errors": parsed_obj.errors}
    else:
        parsed = parse_local(raw)
    if parsed["errors"] and not parsed["entries"]:
        st.error("\n".join(parsed["errors"]))
        return
    if parsed["errors"]:
        st.warning("\n".join(parsed["errors"]))
    enriched = [enrich_entry_from_roster(item, team) for item in parsed["entries"]]
    for item in enriched:
        item["work_date"] = work_date
        item["team"] = team
    st.success(f"특근 {len(enriched)}명을 읽었습니다.")
    st.dataframe(
        [
            {
                "성명": item["name"],
                "회사": item.get("company") or "-",
                "고용형태": item.get("employment_type") or "-",
                "시간": item.get("work_hours"),
                "식수": item.get("meal_count"),
            }
            for item in enriched
        ],
        hide_index=True,
        width="stretch",
    )
    if not can_edit:
        st.warning(reason)
        return
    sig = hashlib.sha256(raw).hexdigest()
    sig_key = f"app_team_xlsx_sig_{survey_id}_{team}_{work_date}"
    if st.session_state.get(sig_key) == sig:
        return
    others = [item for item in existing if str(item.get("work_date") or "") != work_date]
    try:
        replace_team_entries(username, survey_id, team, others + enriched)
        st.session_state[sig_key] = sig
        for key in list(st.session_state.keys()):
            text = str(key)
            if text.startswith((f"cal_roster_{survey_id}_{team}_{work_date}", f"cal_manual_{survey_id}_{team}_{work_date}")):
                del st.session_state[key]
        st.success(f"{work_date} 특근 {len(enriched)}명을 반영했습니다.")
        st.rerun()
    except AccessDenied as exc:
        st.error(str(exc))


def _company_key_app(value: object) -> str:
    text = str(value or "").strip()
    return "" if text in {"", "-", "미지정"} else text


def _render_factory_overtime_list(
    username: str,
    team,
    survey: dict,
    work_date: str,
    rows: list,
    read_only: bool,
) -> None:
    """공장 전체 명단. 우리 팀 인원만 체크 후 삭제한다."""
    from src.store import AccessDenied, list_entries, replace_team_entries, survey_edit_status
    from src.views.calendar_page import _team_entries

    overtime = [item for item in rows if float(item.get("work_hours") or 0) > 0]
    if not overtime:
        st.caption("아직 특근 인원이 없습니다.")
        return
    table = [
        {
            "삭제": False,
            "팀": item.get("team") or "",
            "회사": item.get("company") or "-",
            "성명": item.get("name") or "",
            "고용형태": item.get("employment_type") or "-",
            "시간": item.get("work_hours"),
            "식수": item.get("meal_count") if item.get("meal_count") is not None else "-",
        }
        for item in overtime
    ]
    own_rows = [item for item in overtime if team and str(item.get("team") or "") == team]
    if read_only or not team or not own_rows:
        st.dataframe(
            [{k: v for k, v in row.items() if k != "삭제"} for row in table],
            hide_index=True,
            width="stretch",
        )
        if team and not read_only and not own_rows:
            st.caption("우리 팀 인원이 없어 여기서 삭제할 항목이 없습니다.")
        return

    survey_id = int(survey["id"])
    can_edit, reason = survey_edit_status(username, survey_id)
    edited = st.data_editor(
        table,
        hide_index=True,
        width="stretch",
        disabled=["팀", "회사", "성명", "고용형태", "시간", "식수"],
        column_config={"삭제": st.column_config.CheckboxColumn("삭제", default=False)},
        key=f"cal_del_{survey_id}_{team}_{work_date}",
    )
    st.caption("우리 팀 인원만 체크한 뒤 삭제할 수 있습니다. 잘못 올린 특근을 여기서 지웁니다.")
    if not can_edit:
        st.warning(reason)
        return
    if not st.button("선택한 인원 삭제", key=f"cal_del_btn_{survey_id}_{team}_{work_date}"):
        return
    records = edited.to_dict("records") if hasattr(edited, "to_dict") else list(edited)
    picked = [row for row in records if bool(row.get("삭제"))]
    own_picked = [row for row in picked if str(row.get("팀") or "") == team]
    other_picked = [row for row in picked if str(row.get("팀") or "") != team]
    if other_picked:
        st.error("다른 팀 인원은 삭제할 수 없습니다. 우리 팀만 체크해 주세요.")
        return
    if not own_picked:
        st.warning("삭제할 우리 팀 인원을 체크하세요.")
        return
    remove = {
        (str(row.get("성명") or "").strip(), _company_key_app(row.get("회사")))
        for row in own_picked
    }
    existing = _team_entries(list_entries(username, survey_id), team)
    kept = []
    for item in existing:
        key = (str(item.get("name") or "").strip(), _company_key_app(item.get("company")))
        if str(item.get("work_date") or "") == work_date and key in remove:
            continue
        kept.append(item)
    try:
        replace_team_entries(username, survey_id, team, kept)
        for key in list(st.session_state.keys()):
            text = str(key)
            if text.startswith(
                (
                    f"cal_del_{survey_id}_{team}_{work_date}",
                    f"cal_roster_{survey_id}_{team}_{work_date}",
                    f"cal_manual_{survey_id}_{team}_{work_date}",
                )
            ):
                del st.session_state[key]
        st.success(f"{len(own_picked)}명을 삭제했습니다.")
        st.rerun()
    except AccessDenied as exc:
        st.error(str(exc))


def _is_red_day(day) -> bool:
    try:
        from src.holidays import is_red_day

        return is_red_day(day)
    except Exception:
        pass
    from datetime import date, timedelta

    lunar = {
        2026: ((2, 17), (9, 25), (5, 24)),
        2027: ((2, 7), (9, 15), (5, 13)),
        2028: ((1, 27), (10, 3), (5, 2)),
        2029: ((2, 13), (9, 22), (5, 20)),
        2030: ((2, 3), (9, 12), (5, 9)),
        2031: ((1, 23), (10, 1), (5, 28)),
    }
    elections = {2026: ((6, 3),), 2027: ((3, 3),), 2028: ((4, 12),)}
    solar = (
        (1, 1), (3, 1), (5, 1), (5, 5), (6, 6),
        (7, 17), (8, 15), (10, 3), (10, 9), (12, 25),
    )
    sat_sun_sub = {(3, 1), (5, 1), (5, 5), (7, 17), (8, 15), (10, 3), (10, 9), (12, 25)}
    if day.weekday() == 6:
        return True
    if day.year < 2026 or day.year > 2031:
        return (day.month, day.day) in solar

    holidays = {date(day.year, month, number) for month, number in solar}
    seollal_days = []
    chuseok_days = []
    buddha = None
    table = lunar.get(day.year)
    if table:
        seollal = date(day.year, table[0][0], table[0][1])
        chuseok = date(day.year, table[1][0], table[1][1])
        buddha = date(day.year, table[2][0], table[2][1])
        seollal_days = [seollal + timedelta(days=delta) for delta in (-1, 0, 1)]
        chuseok_days = [chuseok + timedelta(days=delta) for delta in (-1, 0, 1)]
        holidays.update(seollal_days)
        holidays.update(chuseok_days)
        holidays.add(buddha)
    holidays.update(date(day.year, month, number) for month, number in elections.get(day.year, ()))

    def closed(value, blocked):
        return value.weekday() == 6 or value in blocked

    def next_open(value, blocked):
        cursor = value + timedelta(days=1)
        while closed(cursor, blocked):
            cursor += timedelta(days=1)
        if cursor.weekday() == 5:
            cursor += timedelta(days=1)
            while closed(cursor, blocked):
                cursor += timedelta(days=1)
        return cursor

    named_other = {date(day.year, month, number) for month, number in solar}
    named_other.update(date(day.year, month, number) for month, number in elections.get(day.year, ()))
    if buddha is not None:
        named_other.add(buddha)
    triggers = []
    for item in holidays:
        if (item.month, item.day) in sat_sun_sub and item.weekday() >= 5:
            triggers.append(item)
    if buddha is not None and buddha.weekday() >= 5:
        triggers.append(buddha)
    for cluster in (seollal_days, chuseok_days):
        if cluster and (
            any(item.weekday() == 6 for item in cluster)
            or any(item in named_other for item in cluster)
        ):
            triggers.append(max(cluster))
    extras = set()
    for item in sorted(set(triggers)):
        extras.add(next_open(item, holidays | extras))
    return day in holidays | extras


def _render_overtime_calendar(username: str, team=None, read_only: bool = False, **kwargs) -> None:
    from datetime import date

    from src.config import CAFETERIA_MIN_HEADCOUNT
    from src.store import (
        list_overtime_people,
        ensure_open_overtime_survey,
    )
    from src.views.calendar_page import (
        _counts_by_date,
        _month_start,
        _render_team_date_editor,
        _sunday_first_weeks,
    )

    def meal_value(item: dict) -> int:
        try:
            return int(float(item.get("meal_count") or 0))
        except (TypeError, ValueError):
            return 0

    def cell_html(cell_day, day_num: int, dow: int, headcount: int, meals: int, selected: bool) -> str:
        css = "yi-sun" if _is_red_day(cell_day) else ("yi-sat" if dow == 6 else "yi-day")
        on = " yi-cell-on" if selected else ""
        if headcount <= 0:
            meta = (
                "<span class='yi-tag yi-tag-spacer'>&nbsp;</span>"
                "<span class='yi-tag yi-tag-spacer'>&nbsp;</span>"
                "<span class='yi-tag yi-tag-spacer'>&nbsp;</span>"
            )
        else:
            line1 = f"<span class='yi-tag yi-tag-count'>특근 {headcount}명</span>"
            line2 = f"<span class='yi-tag yi-tag-meal'>식수 {meals}명</span>"
            line3 = (
                "<span class='yi-tag yi-tag-cafe'>식당운영</span>"
                if meals >= CAFETERIA_MIN_HEADCOUNT
                else "<span class='yi-tag yi-tag-off'>식당미운영</span>"
            )
            meta = f"{line1}{line2}{line3}"
        return (
            f"<div class='yi-cell{on}'>"
            f"<div class='{css} yi-cell-num'>{day_num}</div>"
            f"<div class='yi-cell-meta'>{meta}</div>"
            "</div>"
        )

    st.header("특근인원")
    st.caption(
        "날짜를 누르면 그날 공장 전체 특근 명단을 볼 수 있습니다. "
        f"입력은 우리 팀만 가능합니다. 달력에는 특근인원과 식수인원을 따로 표시하며, 식수 {CAFETERIA_MIN_HEADCOUNT}명 이상이면 식당을 운영합니다."
    )
    today = date.today()
    if "cal_month" not in st.session_state:
        st.session_state["cal_month"] = _month_start(today)
    month_cursor = st.session_state["cal_month"]
    _render_calendar_month_nav(month_cursor)

    survey = ensure_open_overtime_survey(username)
    all_people = list_overtime_people(username)
    factory_counts: dict[str, int] = {}
    factory_meals: dict[str, int] = {}
    for item in all_people:
        key = str(item.get("work_date") or "")
        if not key:
            continue
        factory_counts[key] = factory_counts.get(key, 0) + 1
        factory_meals[key] = factory_meals.get(key, 0) + meal_value(item)
    team_counts = _counts_by_date(
        [item for item in all_people if team and str(item.get("team") or "") == team]
    )
    left, right = st.columns([7, 5], gap="large")
    with left:
        st.markdown('<div class="yi-card">', unsafe_allow_html=True)
        headers = ["일", "월", "화", "수", "목", "금", "토"]
        head_cols = st.columns(7)
        for index, label in enumerate(headers):
            css = "yi-sun" if index == 0 else ("yi-sat" if index == 6 else "yi-day")
            head_cols[index].markdown(
                f"<div class='{css}' style='text-align:center'>{label}</div>",
                unsafe_allow_html=True,
            )
        selected = st.session_state.get("selected_date")
        for week in _sunday_first_weeks(month_cursor.year, month_cursor.month):
            cols = st.columns(7)
            for dow, day_num in enumerate(week):
                with cols[dow]:
                    if day_num == 0:
                        st.write("")
                        continue
                    day = date(month_cursor.year, month_cursor.month, day_num)
                    iso = day.isoformat()
                    factory_count = factory_counts.get(iso, 0)
                    meal_count = factory_meals.get(iso, 0)
                    st.markdown(
                        cell_html(day, day_num, dow, factory_count, meal_count, selected == iso),
                        unsafe_allow_html=True,
                    )
                    try:
                        clicked = st.button(
                            "열기",
                            key=f"cal_day_{iso}",
                            type="primary" if selected == iso else "secondary",
                            width="stretch",
                        )
                    except TypeError:
                        clicked = st.button(
                            "열기",
                            key=f"cal_day_{iso}",
                            type="primary" if selected == iso else "secondary",
                            use_container_width=True,
                        )
                    if clicked:
                        st.session_state["selected_date"] = iso
                        st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    with right:
        st.markdown('<div class="yi-card">', unsafe_allow_html=True)
        if not st.session_state.get("selected_date"):
            st.info("달력에서 특근일을 선택하세요. 토요일은 파란색, 공휴일은 빨간색입니다.")
            st.markdown("</div>", unsafe_allow_html=True)
            return
        selected_day = date.fromisoformat(str(st.session_state["selected_date"]))
        weekday = "월화수목금토일"[selected_day.weekday()]
        st.subheader(f"{selected_day.month}/{selected_day.day} ({weekday})")
        factory_count = factory_counts.get(selected_day.isoformat(), 0)
        meal_count = factory_meals.get(selected_day.isoformat(), 0)
        team_count = team_counts.get(selected_day.isoformat(), 0)
        cafe = meal_count >= CAFETERIA_MIN_HEADCOUNT
        st.markdown(
            "<div class='yi-summary'>"
            f"<div class='yi-metric'><span>특근인원</span><b>{factory_count}명</b></div>"
            f"<div class='yi-metric'><span>식수인원</span><b>{meal_count}명</b></div>"
            f"<div class='yi-metric'><span>식당</span><b>{'운영' if cafe else '미운영'}</b></div>"
            f"<div class='yi-metric'><span>우리 팀</span><b>{(str(team_count) + '명') if team else '-'}</b></div>"
            "</div>",
            unsafe_allow_html=True,
        )
        if cafe:
            st.success(f"특근 {factory_count}명 · 식수 {meal_count}명 · 식당 운영 (식수 {CAFETERIA_MIN_HEADCOUNT}명 이상)")
        else:
            st.caption(f"식당은 식수인원 {CAFETERIA_MIN_HEADCOUNT}명부터 운영합니다.")
        day_rows = list_overtime_people(username, selected_day.isoformat())
        by_company: dict[str, int] = {}
        by_team: dict[str, int] = {}
        for item in day_rows:
            company = str(item.get("company") or "미지정")
            by_company[company] = by_company.get(company, 0) + 1
            team_name = str(item.get("team") or "미지정")
            by_team[team_name] = by_team.get(team_name, 0) + 1
        if by_team:
            chips = "".join(f"<span class='yi-chip'>{name} {count}명</span>" for name, count in by_team.items())
            st.markdown(chips, unsafe_allow_html=True)
        elif by_company:
            chips = "".join(f"<span class='yi-chip'>{name} {count}명</span>" for name, count in by_company.items())
            st.markdown(chips, unsafe_allow_html=True)
        st.markdown("**공장 전체 특근 명단**")
        _render_factory_overtime_list(
            username, team, survey, selected_day.isoformat(), day_rows, read_only
        )
        if read_only:
            st.caption("공장장은 조회만 가능합니다.")
        elif team:
            st.divider()
            st.markdown(f"**{team} 입력**")
            st.caption("아래 명단만 저장됩니다. 위 전체 명단은 모든 팀이 볼 수 있습니다.")
            _render_team_date_editor(username, team, survey, selected_day.isoformat())
            with st.expander("엑셀로 올리기 · 받기"):
                _render_team_day_excel_io(username, team, survey, selected_day.isoformat())
        st.markdown("</div>", unsafe_allow_html=True)


def _install_calendar_nav_patch() -> None:
    import src.views.calendar_page as calendar_page

    wrapped = _render_overtime_calendar
    wrapped._yi_cal_nav_patched = True
    calendar_page.render_overtime_calendar = wrapped
    try:
        import src.views.shell as shell

        shell.render_overtime_calendar = wrapped
    except Exception:
        pass


def _install_admin_roster_patch() -> None:
    try:
        import src.excel_io as excel_io
        from src.excel_io import RosterParseResult

        def patched_parse(data):
            raw = data if isinstance(data, bytes) else data.getvalue()
            parsed = _parse_roster_bytes(raw)
            return RosterParseResult(rows=parsed["rows"], errors=parsed["errors"])

        excel_io.parse_employee_roster = patched_parse
    except Exception:
        pass

    try:
        import src.views.roster as roster_mod

        def admin_roster(username: str) -> None:
            _render_prodadmin_excel_roster(username)
            st.divider()
            roster_mod.render_team_roster_editor(username, allow_all_teams=True)

        roster_mod._render_admin_roster = admin_roster
    except Exception:
        pass

    import src.views.roster_edit as roster_edit

    original = roster_edit.render_team_roster_editor
    if getattr(original, "_yi_excel_patched", False):
        return

    def wrapped(username, allow_all_teams=False, **kwargs):
        from src.config import ROLE_ADMIN, primary_role

        if primary_role(username) == ROLE_ADMIN and not st.session_state.get("_yi_excel_roster_shown"):
            _render_prodadmin_excel_roster(username)
            st.session_state["_yi_excel_roster_shown"] = True
            st.divider()
        try:
            return original(username, allow_all_teams=allow_all_teams, **kwargs)
        except TypeError:
            return original(username)

    wrapped._yi_excel_patched = True
    roster_edit.render_team_roster_editor = wrapped


def _render_app(authenticator) -> None:
    st.session_state.pop("_yi_excel_roster_shown", None)
    _install_calendar_nav_patch()
    _install_admin_roster_patch()
    from src.views.shell import render_signed_in

    _inject_css(_PAGE_TEXT_CSS)
    username = st.session_state.get("username")
    if not username or username not in USERS:
        st.error("로그인 정보를 확인할 수 없습니다. 다시 로그인하세요.")
        return
    render_signed_in(username, authenticator)


def main() -> None:
    if not _secrets_ready():
        return
    authenticator = _authenticator()
    _restore_cookie(authenticator)
    if st.session_state.get("authentication_status"):
        _render_app(authenticator)
        return
    _render_login(authenticator)


if __name__ == "__main__":
    main()
