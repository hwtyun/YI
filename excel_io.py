"""엑셀 업로드·다운로드. 임직원 명부와 현행 특근 양식을 모두 처리한다."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.config import (
    COMPANIES,
    EMPLOYMENT_TYPES,
    RANK_WORDS,
    SUBMITTING_TEAMS,
    normalize_company,
    normalize_employment,
    normalize_team,
)
from src.hours import parse_work_hours

WEEKDAYS_KR = "월화수목금토일"
DATE_IN_TEXT = re.compile(r"(?:(\d{4})[./-])?(\d{1,2})[./-](\d{1,2})")
YEAR_IN_TEXT = re.compile(r"(20\d{2})")
PLACEHOLDER_YEARS = {1899, 1900, 1904, 2000}
SKIP_ROW_MARKERS = ("특근인원합계", "식수인원합계", "합계")
HEADER_TEAM_MARKERS = {"팀", "소속"}
HEADER_NO_MARKERS = {"no", "n0", "번호", "순번"}
ROSTER_NAME_HEADERS = {"성명", "이름", "name"}
ROSTER_COMPANY_HEADERS = {"회사", "계열사", "소속회사", "company"}
ROSTER_TEAM_HEADERS = {"팀", "소속팀", "소속", "team"}
ROSTER_TYPE_HEADERS = {"고용형태", "고용", "구분", "employment"}
FLAT_DATE_HEADERS = {"특근일자", "일자", "날짜", "work_date"}


@dataclass
class ParseResult:
    entries: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    file_teams: list[str] = field(default_factory=list)


@dataclass
class RosterParseResult:
    rows: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def format_header_date(value: date) -> str:
    return f"{value.month:02d}/{value.day:02d}({WEEKDAYS_KR[value.weekday()]}요일)"


def sane_default_year(year: int | None) -> int:
    today_year = date.today().year
    if year is None or year < 2018 or year in PLACEHOLDER_YEARS:
        return today_year
    return year


def _apply_year(year: int, month: int, day: int, default_year: int) -> str | None:
    use_year = default_year if year in PLACEHOLDER_YEARS or year < 2018 else year
    try:
        return date(use_year, month, day).isoformat()
    except ValueError:
        return None


def parse_header_date(value: object, default_year: int) -> str | None:
    default_year = sane_default_year(default_year)
    if value is None:
        return None
    if isinstance(value, datetime):
        return _apply_year(value.year, value.month, value.day, default_year)
    if isinstance(value, date):
        return _apply_year(value.year, value.month, value.day, default_year)
    text = str(value).strip()
    if not text:
        return None
    match = DATE_IN_TEXT.search(text.replace(" ", ""))
    if not match:
        return None
    year = int(match.group(1)) if match.group(1) else default_year
    month = int(match.group(2))
    day = int(match.group(3))
    return _apply_year(year, month, day, default_year)


def infer_year_from_sheet(sheet: Worksheet) -> int | None:
    for row in sheet.iter_rows(min_row=1, max_row=8, max_col=8, values_only=True):
        for value in row:
            if value is None:
                continue
            if isinstance(value, datetime) and value.year >= 2018:
                return value.year
            if isinstance(value, date) and value.year >= 2018:
                return value.year
            match = YEAR_IN_TEXT.search(str(value))
            if match:
                return int(match.group(1))
    return None


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    cell = ws.cell(row, col)
    if cell.value is not None:
        return cell.value
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).value
    return None


def _seq_no(value: object) -> int | None:
    text = _text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return None


def _text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_rank(value: str) -> bool:
    text = value.strip()
    if not text:
        return False
    if text in RANK_WORDS:
        return True
    return text.endswith("직") and len(text) <= 4


def split_name_rank(col_c: object, col_d: object) -> tuple[str, str]:
    """헤더는 직급|성명이지만 실무는 C=성명, D=직급인 경우가 많다."""
    left = _text(col_c)
    right = _text(col_d)
    left_rank = _looks_like_rank(left)
    right_rank = _looks_like_rank(right)
    if left_rank and not right_rank:
        return right, left
    if right_rank and not left_rank:
        return left, right
    return left, right


def _header_map(values: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, raw in enumerate(values):
        key = _text(raw).replace(" ", "").lower()
        if key:
            mapping[key] = index
    return mapping


def _find_col(headers: dict[str, int], aliases: set[str]) -> int | None:
    folded = {item.replace(" ", "").lower() for item in aliases}
    for key, index in headers.items():
        if key in folded:
            return index
    return None


def _is_skip_row(team_text: str, name: str, hours_raw: object) -> bool:
    combined = f"{team_text}{name}"
    if any(marker in combined for marker in SKIP_ROW_MARKERS):
        return True
    if "합계" in team_text:
        return True
    return not name and hours_raw in (None, "")


def _looks_like_block_header(values: list[object]) -> bool:
    texts = [_text(item) for item in values]
    folded = [item.replace(" ", "").lower() for item in texts]
    has_team = any(item in HEADER_TEAM_MARKERS for item in texts)
    has_no = any(item in HEADER_NO_MARKERS for item in folded)
    has_person = any(item in ("직급", "성명", "이름") for item in texts)
    return has_team and has_no and has_person


def parse_employee_roster(data: bytes | BytesIO) -> RosterParseResult:
    workbook = load_workbook(filename=BytesIO(data) if isinstance(data, bytes) else data, data_only=True)
    sheet = workbook.active
    header_row = None
    for row in sheet.iter_rows(min_row=1, max_row=20, max_col=12, values_only=True):
        texts = [_text(item) for item in row]
        headers = _header_map(texts)
        if _find_col(headers, ROSTER_NAME_HEADERS) is not None and _find_col(headers, ROSTER_TEAM_HEADERS) is not None:
            header_row = texts
            break
    result = RosterParseResult()
    if header_row is None:
        result.errors.append("성명·팀 헤더를 찾지 못했습니다. 템플릿 양식을 사용해 주세요.")
        return result

    headers = _header_map(header_row)
    name_col = _find_col(headers, ROSTER_NAME_HEADERS)
    company_col = _find_col(headers, ROSTER_COMPANY_HEADERS)
    team_col = _find_col(headers, ROSTER_TEAM_HEADERS)
    type_col = _find_col(headers, ROSTER_TYPE_HEADERS)
    if name_col is None or company_col is None or team_col is None or type_col is None:
        result.errors.append("필수 열(성명, 회사, 팀, 고용형태)이 없습니다.")
        return result

    seen: set[tuple[str, str, str]] = set()
    header_index = None
    for index, row in enumerate(sheet.iter_rows(min_row=1, max_col=12, values_only=True), start=1):
        texts = [_text(item) for item in row]
        if texts == header_row:
            header_index = index
            break
    start = (header_index or 1) + 1
    for excel_row, row in enumerate(
        sheet.iter_rows(min_row=start, max_col=12, values_only=True),
        start=start,
    ):
        values = list(row)
        name = _text(values[name_col] if name_col < len(values) else "")
        if not name or name == "안내" or name.startswith(("회사는", "팀은", "고용형태는", "자주", "고정")):
            continue
        company = normalize_company(values[company_col] if company_col < len(values) else "")
        team = normalize_team(values[team_col] if team_col < len(values) else "")
        employment = normalize_employment(values[type_col] if type_col < len(values) else "") or "정규직"
        if company is None:
            result.errors.append(f"{excel_row}행 '{name}': 회사는 에이텍모빌리티 또는 에이텍컴퓨터여야 합니다.")
            continue
        if team is None:
            result.errors.append(f"{excel_row}행 '{name}': 팀을 확인할 수 없습니다.")
            continue
        key = (name, company, team)
        if key in seen:
            result.errors.append(f"{excel_row}행 '{name}': 같은 회사·팀에 이름이 중복됩니다.")
            continue
        seen.add(key)
        result.rows.append(
            {
                "name": name,
                "company": company,
                "team": team,
                "employment_type": employment,
            }
        )
    if not result.rows and not result.errors:
        result.errors.append("유효한 임직원 행이 없습니다.")
    return result


def build_employee_template(rows: list[dict[str, Any]] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "임직원명부"
    headers = ["성명", "회사", "팀", "고용형태"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    if rows is None:
        data_rows = [
            {
                "name": "홍길동",
                "company": COMPANIES[0],
                "team": "제조팀",
                "employment_type": "정규직",
            },
            {
                "name": "김고정",
                "company": COMPANIES[1],
                "team": "제조팀",
                "employment_type": "일용직",
            },
        ]
    else:
        data_rows = rows
    for item in data_rows:
        sheet.append(
            [
                item.get("name") or "",
                item.get("company") or "",
                item.get("team") or "",
                item.get("employment_type") or "",
            ]
        )
    sheet.append([])
    sheet.append(["안내"])
    sheet.append(["회사는 에이텍모빌리티 또는 에이텍컴퓨터만 입력합니다."])
    sheet.append([f"팀은 {', '.join(SUBMITTING_TEAMS)} 중 하나입니다."])
    sheet.append(["고용형태는 정규직, 계약직, 일용직입니다. 책임·선임·사원 같은 직급을 적어도 정규직으로 저장합니다."])
    sheet.append(["자주 바뀌는 일용직은 명부에 넣지 않아도 됩니다. 팀은 화면에서 수기로 추가합니다."])
    sheet.append(["고정 일용직만 명부에 넣으면 해당 팀 입력 화면에 함께 나타납니다."])
    for index, width in enumerate((14, 18, 16, 12), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _is_flat_overtime_sheet(sheet: Worksheet) -> bool:
    first = [_text(item) for item in next(sheet.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True))]
    headers = _header_map(first)
    return (
        _find_col(headers, ROSTER_NAME_HEADERS) is not None
        and _find_col(headers, FLAT_DATE_HEADERS) is not None
    )


def _parse_flat_overtime(sheet: Worksheet, default_year: int) -> ParseResult:
    result = ParseResult()
    header = [_text(item) for item in next(sheet.iter_rows(min_row=1, max_row=1, max_col=16, values_only=True))]
    headers = _header_map(header)
    name_col = _find_col(headers, ROSTER_NAME_HEADERS)
    date_col = _find_col(headers, FLAT_DATE_HEADERS)
    team_col = _find_col(headers, ROSTER_TEAM_HEADERS)
    company_col = _find_col(headers, ROSTER_COMPANY_HEADERS)
    type_col = _find_col(headers, ROSTER_TYPE_HEADERS)
    hours_col = _find_col(headers, {"근무시간", "시간", "work_hours"})
    meal_col = _find_col(headers, {"식수인원", "식수", "meal"})
    note_col = _find_col(headers, {"비고", "note"})
    rank_col = _find_col(headers, {"직급", "rank"})
    manual_col = _find_col(headers, {"수기", "일용직수기"})
    teams: set[str] = set()
    for row in sheet.iter_rows(min_row=2, max_col=16, values_only=True):
        values = list(row)
        name = _text(values[name_col] if name_col is not None and name_col < len(values) else "")
        work_date = parse_header_date(
            values[date_col] if date_col is not None and date_col < len(values) else None,
            default_year,
        )
        if not name or not work_date:
            continue
        team_raw = values[team_col] if team_col is not None and team_col < len(values) else ""
        team = normalize_team(team_raw) or _text(team_raw)
        if team:
            teams.add(team)
        hours_raw = values[hours_col] if hours_col is not None and hours_col < len(values) else 0
        meal_raw = values[meal_col] if meal_col is not None and meal_col < len(values) else None
        company_raw = values[company_col] if company_col is not None and company_col < len(values) else ""
        type_raw = values[type_col] if type_col is not None and type_col < len(values) else ""
        rank = _text(values[rank_col] if rank_col is not None and rank_col < len(values) else "") or None
        note = _text(values[note_col] if note_col is not None and note_col < len(values) else "") or None
        manual_raw = values[manual_col] if manual_col is not None and manual_col < len(values) else ""
        result.entries.append(
            {
                "team": team or None,
                "rank": rank,
                "name": name,
                "work_date": work_date,
                "work_hours": parse_work_hours(hours_raw),
                "meal_count": None if meal_raw in (None, "") else parse_work_hours(meal_raw),
                "note": note,
                "company": normalize_company(company_raw),
                "employment_type": normalize_employment(type_raw),
                "is_manual": 1 if _text(manual_raw) in {"1", "Y", "y", "예", "수기", "TRUE", "True"} else 0,
            }
        )
    result.file_teams = sorted(teams)
    return result


def _parse_block_overtime(sheet: Worksheet, default_year: int) -> ParseResult:
    result = ParseResult()
    max_row = sheet.max_row or 1
    max_col = max(sheet.max_column or 7, 7)
    current_date: str | None = None
    last_team = ""
    data_started = False
    teams: set[str] = set()
    row = 1
    while row <= max_row:
        values = [_cell_value(sheet, row, col) for col in range(1, max_col + 1)]
        if _looks_like_block_header(values):
            current_date = None
            for value in values[4:]:
                parsed = parse_header_date(value, default_year)
                if parsed:
                    current_date = parsed
                    break
            next_values = [_cell_value(sheet, row + 1, col) for col in range(1, max_col + 1)]
            next_texts = [_text(item) for item in next_values]
            if "근무시간" in next_texts or "식수인원" in next_texts:
                row += 2
            else:
                row += 1
            data_started = True
            last_team = ""
            if current_date is None:
                result.warnings.append(f"{row}행 근처 헤더에서 특근일자를 읽지 못했습니다.")
            continue

        if not data_started:
            row += 1
            continue

        team_text = _text(values[0]) or last_team
        if team_text:
            last_team = team_text
        name, rank = split_name_rank(values[2] if len(values) > 2 else "", values[3] if len(values) > 3 else "")
        hours_raw = values[4] if len(values) > 4 else None
        meal_raw = values[5] if len(values) > 5 else None
        note = _text(values[6] if len(values) > 6 else "") or None
        if _is_skip_row(team_text, name, hours_raw):
            row += 1
            continue
        if current_date is None:
            result.warnings.append(f"{row}행 '{name}': 일자가 없어 건너뛰었습니다.")
            row += 1
            continue
        team = normalize_team(team_text) or team_text
        if team:
            teams.add(team)
        result.entries.append(
            {
                "team": team or None,
                "seq_no": _seq_no(values[1] if len(values) > 1 else None),
                "rank": rank or None,
                "name": name,
                "work_date": current_date,
                "work_hours": parse_work_hours(hours_raw),
                "meal_count": None if meal_raw in (None, "") else parse_work_hours(meal_raw),
                "note": note,
                "company": None,
                "employment_type": None,
                "is_manual": 0,
            }
        )
        row += 1
    result.file_teams = sorted(item for item in teams if item)
    if not result.entries:
        result.warnings.append("인원 행을 찾지 못했습니다. 일자 블록 양식인지 확인해 주세요.")
    return result


def parse_overtime_workbook(data: bytes | BytesIO, default_year: int | None = None) -> ParseResult:
    workbook = load_workbook(filename=BytesIO(data) if isinstance(data, bytes) else data, data_only=True)
    fallback_year = sane_default_year(default_year)
    combined = ParseResult()
    for sheet in workbook.worksheets:
        year = infer_year_from_sheet(sheet) or fallback_year
        parsed = _parse_flat_overtime(sheet, year) if _is_flat_overtime_sheet(sheet) else _parse_block_overtime(sheet, year)
        combined.entries.extend(parsed.entries)
        combined.warnings.extend(parsed.warnings)
        for team in parsed.file_teams:
            if team not in combined.file_teams:
                combined.file_teams.append(team)
    return combined


def filter_entries_for_team(
    parsed: ParseResult,
    login_team: str,
) -> ParseResult:
    """로그인 팀을 우선한다. 파일 팀명이 다르면 경고만 남기고 본인 팀으로 저장한다."""
    result = ParseResult(file_teams=list(parsed.file_teams), warnings=list(parsed.warnings))
    others = [team for team in parsed.file_teams if team and team != login_team]
    if others:
        result.warnings.append(
            f"파일의 팀명({', '.join(others)})이 로그인 팀({login_team})과 다릅니다. 로그인 팀으로 저장합니다."
        )
    for item in parsed.entries:
        row = dict(item)
        row["team"] = login_team
        result.entries.append(row)
    return result


def build_overtime_workbook(
    title: str,
    entries: list[dict[str, Any]],
    include_source_team: bool = False,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "특근인원"
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    dates = sorted({str(item.get("work_date")) for item in entries if item.get("work_date")})
    sheet.merge_cells("A1:G2")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = center
    row = 4
    if not dates:
        sheet["A4"] = "저장된 인원이 없습니다."
    for work_date in dates:
        day = date.fromisoformat(work_date)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)
        sheet.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)
        sheet.merge_cells(start_row=row, start_column=4, end_row=row + 1, end_column=4)
        sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        sheet.merge_cells(start_row=row, start_column=7, end_row=row + 1, end_column=7)
        headers = ["팀", "NO", "직급", "성명", format_header_date(day), None, "비고"]
        for col, value in enumerate(headers, start=1):
            if value is None:
                continue
            cell = sheet.cell(row, col, value)
            cell.font = Font(bold=True)
            cell.alignment = center
            cell.border = thin
        sheet.cell(row + 1, 5, "근무시간").alignment = center
        sheet.cell(row + 1, 5).border = thin
        sheet.cell(row + 1, 6, "식수인원").alignment = center
        sheet.cell(row + 1, 6).border = thin
        for col in (1, 2, 3, 4, 7):
            sheet.cell(row + 1, col).border = thin
        row += 2
        date_rows = [item for item in entries if str(item.get("work_date")) == work_date]
        date_rows.sort(key=lambda item: (str(item.get("team") or ""), int(item.get("seq_no") or 0), str(item.get("name") or "")))
        start_data = row
        last_team = None
        team_start = row
        seq = 0
        for item in date_rows:
            team = str(item.get("team") or "")
            if team != last_team:
                if last_team is not None and row - 1 >= team_start:
                    sheet.merge_cells(start_row=team_start, start_column=1, end_row=row - 1, end_column=1)
                last_team = team
                team_start = row
                seq = 0
            seq += 1
            hours = parse_work_hours(item.get("work_hours", 0))
            hours_text = f"{int(hours)}H" if hours == int(hours) else f"{hours}H"
            meal = item.get("meal_count")
            values = [
                team,
                seq,
                item.get("name") or "",
                item.get("rank") or "",
                hours_text,
                "" if meal in (None, "") else meal,
                item.get("note") or "",
            ]
            for col, value in enumerate(values, start=1):
                cell = sheet.cell(row, col, value)
                cell.alignment = center
                cell.border = thin
            row += 1
        if last_team is not None and row - 1 >= team_start:
            sheet.merge_cells(start_row=team_start, start_column=1, end_row=row - 1, end_column=1)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.cell(row, 1, "특근인원합계").alignment = center
        sheet.cell(row, 3, len(date_rows))
        sheet.cell(row, 4, "식수인원")
        if date_rows:
            sheet.cell(row, 6, f"=SUM(F{start_data}:F{row - 1})")
        for col in range(1, 8):
            sheet.cell(row, col).border = thin
            sheet.cell(row, col).alignment = center
        row += 2
    if include_source_team:
        sheet["A3"] = "출처팀은 A열(팀)입니다. 메일은 보내지 않으니 다운로드 후 직접 검토하세요."
    for index, width in enumerate((16, 8, 10, 12, 12, 12, 16), start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_aggregate_workbook(
    title: str,
    overtime: list[dict[str, Any]],
    date_totals: list[Any],
    anomalies: list[Any] | None = None,
    team_status: list[Any] | None = None,
) -> bytes:
    """취합본: 특근인원(0시간 제외) + 일자별 합계 + 이상치 + 제출 현황."""
    data = build_overtime_workbook(title, overtime, include_source_team=True)
    workbook = load_workbook(BytesIO(data))
    total_sheet = workbook.create_sheet("일자별합계")
    total_sheet.append(["특근일자", "특근인원", "식수인원", "근무시간합계"])
    for item in date_totals:
        total_sheet.append([item.work_date, item.headcount, item.meal_sum, item.hours_sum])
    if date_totals:
        total_sheet.append(
            [
                "합계",
                sum(item.headcount for item in date_totals),
                sum(item.meal_sum for item in date_totals),
                sum(item.hours_sum for item in date_totals),
            ]
        )
    for index, width in enumerate((14, 12, 12, 14), start=1):
        total_sheet.column_dimensions[get_column_letter(index)].width = width

    if team_status is not None:
        status_sheet = workbook.create_sheet("제출현황")
        status_sheet.append(["팀", "상태", "제출시각", "저장인원", "특근인원"])
        for item in team_status:
            status_sheet.append(
                [
                    item.team,
                    item.label,
                    item.submitted_at or "",
                    item.saved_count,
                    item.overtime_count,
                ]
            )
        for index, width in enumerate((16, 12, 24, 12, 12), start=1):
            status_sheet.column_dimensions[get_column_letter(index)].width = width

    issue_sheet = workbook.create_sheet("이상치")
    issue_sheet.append(["수준", "구분", "내용", "일자", "성명", "팀"])
    for item in anomalies or []:
        issue_sheet.append(
            [
                "오류" if item.level == "error" else "확인",
                item.kind,
                item.message,
                item.work_date or "",
                item.name or "",
                item.team or "",
            ]
        )
    if not anomalies:
        issue_sheet.append(["", "", "표시할 이상치가 없습니다.", "", "", ""])
    for index, width in enumerate((10, 16, 50, 14, 12, 16), start=1):
        issue_sheet.column_dimensions[get_column_letter(index)].width = width

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_generic_workbook(
    title: str,
    schema: dict[str, Any],
    rows: list[dict[str, Any]],
    anomalies: list[Any] | None = None,
    team_status: list[Any] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "취합"
    columns = list(schema.get("columns") or [])
    headers = ["출처팀"] + [str(item.get("label") or item.get("key")) for item in columns]
    sheet.append(headers)
    keys = [str(item.get("key")) for item in columns]
    for row in rows:
        sheet.append(
            [row.get("source_team") or row.get("team") or ""]
            + ["" if row.get(key) is None else row.get(key) for key in keys]
        )
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.append([])
    sheet.append([title])
    sheet.append(["메일은 보내지 않으니 다운로드 후 직접 검토하세요."])
    for index in range(1, max(len(headers), 1) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 16

    if team_status is not None:
        status_sheet = workbook.create_sheet("제출현황")
        status_sheet.append(["팀", "상태", "제출시각", "저장행", "입력행"])
        for item in team_status:
            status_sheet.append(
                [item.team, item.label, item.submitted_at or "", item.saved_count, item.overtime_count]
            )

    issue_sheet = workbook.create_sheet("이상치")
    issue_sheet.append(["수준", "구분", "내용", "팀"])
    for item in anomalies or []:
        issue_sheet.append(
            [
                "오류" if item.level == "error" else "확인",
                item.kind,
                item.message,
                item.team or "",
            ]
        )
    if not anomalies:
        issue_sheet.append(["", "", "표시할 이상치가 없습니다.", ""])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()

